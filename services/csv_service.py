"""
CSV / TSV / 테이블 데이터 뷰어 백엔드 서비스 모듈
- 인코딩 자동 감지 (UTF-8, UTF-8-BOM, CP949/EUC-KR, Latin-1)
- 구분자 자동 감지 (쉼표, 탭, 세미콜론, 파이프)
- 대화상자 파일 열기 및 저장 지원
"""
import os
import csv
import io
import eel


def _detect_encoding_and_read(file_path):
    """파일의 인코딩(UTF-8, CP949 등)을 안전하게 감지하여 텍스트 및 인코딩명 반환"""
    encodings_to_try = ['utf-8-sig', 'utf-8', 'cp949', 'euc-kr', 'latin-1']
    
    with open(file_path, 'rb') as f:
        raw_bytes = f.read()

    for enc in encodings_to_try:
        try:
            text = raw_bytes.decode(enc)
            return text, enc
        except UnicodeDecodeError:
            continue

    return raw_bytes.decode('utf-8', errors='replace'), 'utf-8 (fallback)'


def _detect_delimiter(sample_text):
    """텍스트 샘플을 기반으로 가장 적합한 구분자 판별"""
    if not sample_text:
        return ','
    
    lines = [line.strip() for line in sample_text.splitlines()[:15] if line.strip()]
    if not lines:
        return ','

    sample_chunk = '\n'.join(lines)
    try:
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(sample_chunk, delimiters=[',', '\t', ';', '|'])
        return dialect.delimiter
    except Exception:
        pass

    delimiters = [',', '\t', ';', '|']
    counts = {d: sum(line.count(d) for line in lines) for d in delimiters}
    best_delimiter = max(counts, key=counts.get)
    return best_delimiter if counts[best_delimiter] > 0 else ','


def _parse_csv_lines(text, delimiter=None):
    """문자열 텍스트를 파싱하여 headers와 rows 2차원 리스트 반환"""
    if not text or not text.strip():
        return [], [], ','

    if not delimiter or delimiter == 'auto':
        delimiter = _detect_delimiter(text)

    f = io.StringIO(text)
    reader = csv.reader(f, delimiter=delimiter)
    
    all_rows = []
    for row in reader:
        if not row or all(cell == '' for cell in row):
            continue
        all_rows.append(row)

    if not all_rows:
        return [], [], delimiter

    headers = all_rows[0]
    rows = all_rows[1:]

    max_cols = max(len(headers), max((len(r) for r in rows), default=0))
    if len(headers) < max_cols:
        headers.extend([f'열_{i+1}' for i in range(len(headers), max_cols)])

    normalized_rows = []
    for r in rows:
        if len(r) < max_cols:
            r = r + [''] * (max_cols - len(r))
        elif len(r) > max_cols:
            r = r[:max_cols]
        normalized_rows.append(r)

    return headers, normalized_rows, delimiter


@eel.expose
def select_and_read_csv_file():
    """탐색기 파일 대화상자를 열어 CSV/TSV 파일을 읽고 파싱된 데이터 반환"""
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        file_selected = filedialog.askopenfilename(
            title='열람할 CSV / TSV 데이터 파일 선택',
            filetypes=[
                ('CSV / TSV / 텍스트 파일', '*.csv;*.tsv;*.txt;*.tab;*.dat'),
                ('CSV 파일 (*.csv)', '*.csv'),
                ('TSV / 탭 구분 파일 (*.tsv;*.tab;*.txt)', '*.tsv;*.tab;*.txt'),
                ('모든 파일 (*.*)', '*.*')
            ]
        )
        root.destroy()

        if not file_selected:
            return {'status': 'cancelled'}

        file_selected = os.path.normpath(file_selected)
        return read_csv_from_path(file_selected)

    except Exception as e:
        return {'status': 'error', 'message': f'파일 선택 중 오류 발생: {str(e)}'}


@eel.expose
def read_csv_from_path(file_path, forced_encoding='auto', forced_delimiter='auto'):
    """지정된 파일 경로에서 CSV를 읽어 파싱하여 반환"""
    try:
        if not os.path.exists(file_path):
            return {'status': 'error', 'message': f'파일을 찾을 수 없습니다: {file_path}'}

        file_size = os.path.getsize(file_path)
        file_name = os.path.basename(file_path)

        if forced_encoding and forced_encoding != 'auto':
            try:
                with open(file_path, 'r', encoding=forced_encoding, errors='replace') as f:
                    text = f.read()
                encoding_used = forced_encoding
            except Exception:
                text, encoding_used = _detect_encoding_and_read(file_path)
        else:
            text, encoding_used = _detect_encoding_and_read(file_path)

        delimiter_to_use = None if forced_delimiter == 'auto' else forced_delimiter
        headers, rows, detected_delimiter = _parse_csv_lines(text, delimiter=delimiter_to_use)

        return {
            'status': 'success',
            'file_name': file_name,
            'file_path': file_path,
            'file_size': file_size,
            'encoding': encoding_used,
            'delimiter': detected_delimiter,
            'headers': headers,
            'rows': rows,
            'total_rows': len(rows),
            'total_cols': len(headers)
        }
    except Exception as e:
        return {'status': 'error', 'message': f'CSV 파일 읽기 실패: {str(e)}'}


@eel.expose
def parse_raw_csv_text(raw_text, forced_delimiter='auto'):
    """클립보드 또는 드롭된 원시 텍스트를 CSV로 파싱"""
    try:
        if not raw_text or not raw_text.strip():
            return {'status': 'error', 'message': '파싱할 데이터가 비어 있습니다.'}

        delimiter_to_use = None if forced_delimiter == 'auto' else forced_delimiter
        headers, rows, detected_delimiter = _parse_csv_lines(raw_text, delimiter=delimiter_to_use)

        return {
            'status': 'success',
            'file_name': '클립보드 데이터 (Pasted Data)',
            'file_path': '',
            'file_size': len(raw_text.encode('utf-8')),
            'encoding': 'utf-8',
            'delimiter': detected_delimiter,
            'headers': headers,
            'rows': rows,
            'total_rows': len(rows),
            'total_cols': len(headers)
        }
    except Exception as e:
        return {'status': 'error', 'message': f'텍스트 파싱 실패: {str(e)}'}


@eel.expose
def export_csv_to_excel(headers, rows, default_filename='export.xlsx'):
    """
    현재 테이블 데이터(headers, rows)를 스타일링된 Excel(.xlsx) 파일로 내보내기
    """
    try:
        import tkinter as tk
        from tkinter import filedialog
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if not headers and not rows:
            return {'status': 'error', 'message': '내보낼 데이터가 없습니다.'}

        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        base_name = os.path.splitext(default_filename)[0] if default_filename else 'export'
        save_path = filedialog.asksaveasfilename(
            title='Excel (.xlsx) 파일 저장',
            initialfile=f"{base_name}.xlsx",
            defaultextension='.xlsx',
            filetypes=[
                ('Excel 통합 문서 (*.xlsx)', '*.xlsx'),
                ('모든 파일 (*.*)', '*.*')
            ]
        )
        root.destroy()

        if not save_path:
            return {'status': 'cancelled'}

        save_path = os.path.normpath(save_path)

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.views.sheetView[0].showGridLines = True

        # 스타일 정의
        header_font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        data_font = Font(name='맑은 고딕', size=9.5)
        zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # 1. 헤더 쓰기
        if headers:
            ws.append(headers)
            ws.row_dimensions[1].height = 26
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

        # 2. 본문 데이터 쓰기
        for r_idx, row in enumerate(rows, start=2 if headers else 1):
            ws.append(row)
            ws.row_dimensions[r_idx].height = 20
            is_even = (r_idx % 2 == 0)
            row_fill = zebra_fill if is_even else white_fill

            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = data_font
                cell.fill = row_fill
                cell.border = thin_border

                # 숫자 판별 및 서식 적용
                str_val = str(val).strip()
                if str_val != '' and str_val.replace(',', '').replace('.', '', 1).replace('-', '', 1).isdigit():
                    try:
                        clean_num_str = str_val.replace(',', '')
                        if '.' in clean_num_str:
                            num = float(clean_num_str)
                            cell.value = num
                            cell.number_format = '#,##0.00'
                        else:
                            num = int(clean_num_str)
                            cell.value = num
                            cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    except Exception:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # 3. 열 너비 자동 맞춤
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                if char_len > max_len:
                    max_len = char_len
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 4, 50))

        wb.save(save_path)

        return {
            'status': 'success',
            'path': save_path,
            'file_name': os.path.basename(save_path),
            'total_rows': len(rows),
            'total_cols': len(headers)
        }
    except Exception as e:
        return {'status': 'error', 'message': f'Excel 저장 실패: {str(e)}'}


@eel.expose
def save_csv_to_file(raw_content, default_filename='export.csv'):
    """파일 저장 대화상자를 열어 CSV 텍스트를 파일로 저장 (UTF-8 with BOM for Excel)"""
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        save_path = filedialog.asksaveasfilename(
            title='CSV 파일 저장',
            initialfile=default_filename,
            defaultextension='.csv',
            filetypes=[
                ('CSV 파일 (*.csv)', '*.csv'),
                ('TSV 파일 (*.tsv)', '*.tsv'),
                ('모든 파일 (*.*)', '*.*')
            ]
        )
        root.destroy()

        if not save_path:
            return {'status': 'cancelled'}

        save_path = os.path.normpath(save_path)
        with open(save_path, 'w', encoding='utf-8-sig', newline='') as f:
            f.write(raw_content)

        return {
            'status': 'success',
            'path': save_path,
            'file_name': os.path.basename(save_path)
        }
    except Exception as e:
        return {'status': 'error', 'message': f'파일 저장 실패: {str(e)}'}
