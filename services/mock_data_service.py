"""
모의 데이터(Mock Data) 및 엑셀(.xlsx) 대량 생성 서비스 모듈
- 한국인 이름, 전화번호, 커스텀 도메인 이메일, 사용자 지정 문자열 목록(부서/직급/직책/상태 등)
- openpyxl 기반 스타일 서식 엑셀 파일 생성 및 CSV/JSON 내보내기 지원
"""
import os
import random
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog
import eel

# 한국 성씨 및 이름 데이터베이스
LAST_NAMES = [
    "김", "이", "박", "최", "정", "강", "조", "윤", "장", "임",
    "한", "오", "서", "신", "권", "황", "안", "송", "류", "홍",
    "고", "문", "양", "손", "배", "백", "허", "유", "남", "심"
]

MALE_FIRST_NAMES = [
    "민준", "서준", "도윤", "예준", "시우", "하준", "주원", "지호", "지후", "준우",
    "준서", "건우", "도현", "현우", "지훈", "우진", "선우", "서진", "유준", "연우",
    "민재", "정우", "승우", "승현", "시윤", "준혁", "은우", "지환", "승민", "유찬",
    "동현", "성민", "재원", "태민", "진우", "준영", "현준", "태양", "영호", "성호"
]

FEMALE_FIRST_NAMES = [
    "서연", "서윤", "지우", "서현", "하은", "민서", "지유", "윤서", "채원", "지원",
    "은서", "다은", "예은", "수아", "지아", "소율", "예원", "지원", "예린", "소은",
    "하린", "유진", "시은", "수빈", "채은", "지민", "예나", "다인", "채윤", "서영",
    "나은", "민아", "유나", "아린", "소담", "혜원", "은채", "가은", "예진", "수연"
]

KOREAN_ROMAN_MAP = {
    '김': 'kim', '이': 'lee', '박': 'park', '최': 'choi', '정': 'jung', '강': 'kang',
    '조': 'cho', '윤': 'yoon', '장': 'jang', '임': 'lim', '한': 'han', '오': 'oh',
    '서': 'seo', '신': 'shin', '권': 'kwon', '황': 'hwang', '안': 'ahn', '송': 'song',
    '류': 'ryu', '홍': 'hong', '고': 'ko', '문': 'moon', '양': 'yang', '손': 'sohn',
    '배': 'bae', '백': 'baek', '허': 'heo', '유': 'yoo', '남': 'nam', '심': 'shim',
    '민준': 'minjun', '서준': 'seojun', '도윤': 'doyoon', '예준': 'yejun', '시우': 'siwoo',
    '하준': 'hajun', '서연': 'seoyeon', '서윤': 'seoyoon', '지우': 'jiwoo', '서현': 'seohyun',
    '하은': 'haeun', '민서': 'minseo', '지유': 'jiyu', '윤서': 'yoonseo', '채원': 'chaewon',
    '지원': 'jiwon', '수빈': 'soobin', '태민': 'taemin', '도현': 'dohyun', '준혁': 'junhyeok',
    '현우': 'hyunwoo', '지훈': 'jihoon', '유진': 'yoojin', '지민': 'jimin', '은우': 'eunwoo'
}

KOREA_CITIES = ["서울특별시", "경기도", "인천광역시", "부산광역시", "대구광역시", "대전광역시", "광주광역시", "울산광역시", "세종특별자치시"]
KOREA_DISTRICTS = {
    "서울특별시": ["강남구 테헤란로", "서초구 서초대로", "송파구 올림픽로", "영등포구 여의대로", "마포구 양화로", "중구 세종대로", "성동구 왕십리로"],
    "경기도": ["성남시 분당구 판교역로", "수원시 영통구 광교로", "고양시 일산동구 중앙로", "용인시 수지구 포은대로", "안양시 동안구 시민대로"],
    "인천광역시": ["연수구 송도과학로", "남동구 예술로", "부평구 부평대로", "서구 청라커낼로"],
    "부산광역시": ["해운대구 센텀중앙로", "부산진구 가야대로", "수영구 광안해변로"],
    "대전광역시": ["유성구 대덕대로", "서구 둔산로", "중구 중앙로"]
}

COMPANY_PRESETS = ["(주)테크솔루션", "(주)넥스트이노베이션", "(주)글로벌네트웍스", "(주)쿠첸", "(주)한국디지털", "(주)스마트시스템", "(주)한양코퍼레이션"]
PRODUCTS = [
    ("고효율 IH 압력밥솥 6인용", 328000), ("스마트 인덕션 3구 하이브리드", 590000),
    ("초경량 무선 청소기 V10", 249000), ("스마트 공기청정기 파워케어", 189000),
    ("대용량 에어프라이어 7L", 129000), ("음성인식 가습기 프로", 89000),
    ("프리미엄 믹서기 블렌더", 79000), ("전기포트 스테인리스 1.7L", 45000)
]


def _generate_korean_name(gender="any"):
    last = random.choice(LAST_NAMES)
    if gender == "male":
        first = random.choice(MALE_FIRST_NAMES)
    elif gender == "female":
        first = random.choice(FEMALE_FIRST_NAMES)
    else:
        first = random.choice(MALE_FIRST_NAMES if random.random() < 0.5 else FEMALE_FIRST_NAMES)
    return last + first, last, first


def _generate_email(name, domain_list=None):
    domains = [d.strip() for d in (domain_list or "").split(",") if d.strip()]
    if not domains:
        domains = ["gmail.com", "naver.com", "kakao.com", "daum.net"]
    domain = random.choice(domains)
    if domain.startswith("@"):
        domain = domain[1:]

    # 영문 아이디 생성
    last = name[0] if len(name) > 0 else ""
    first = name[1:] if len(name) > 1 else ""
    r_last = KOREAN_ROMAN_MAP.get(last, "user")
    r_first = KOREAN_ROMAN_MAP.get(first, str(random.randint(100, 999)))

    patterns = [
        f"{r_first}.{r_last}",
        f"{r_last}_{r_first}",
        f"{r_first}{random.randint(1, 99)}",
        f"{r_last[0]}{r_first}",
        f"{r_first}_{random.randint(10, 99)}"
    ]
    user_id = random.choice(patterns).lower()
    return f"{user_id}@{domain}"


def _generate_phone(format_type="010-XXXX-XXXX"):
    mid = random.randint(2000, 9999)
    last = random.randint(1000, 9999)
    if format_type == "010XXXXXXXX":
        return f"010{mid}{last}"
    elif format_type == "02-XXX-XXXX":
        return f"02-{random.randint(200, 899)}-{last}"
    return f"010-{mid}-{last}"


def _generate_biz_no():
    weights = [1, 3, 7, 1, 3, 7, 1, 3, 5]
    digits = [random.randint(1, 9)] + [random.randint(0, 9) for _ in range(7)]
    chk_sum = sum(w * d for w, d in zip(weights, digits))
    p9 = weights[8] * digits[8] if len(digits) > 8 else 0
    digits.append(random.randint(0, 9))
    chk_sum = sum(w * d for w, d in zip(weights[:9], digits[:9]))
    chk_sum += (weights[8] * digits[8]) // 10
    chk_digit = (10 - (chk_sum % 10)) % 10
    digits.append(chk_digit)
    raw = "".join(map(str, digits[:10]))
    return f"{raw[:3]}-{raw[3:5]}-{raw[5:]}"


def _generate_date(start_year=2020, end_year=2026, date_format="%Y-%m-%d"):
    start_dt = datetime.date(int(start_year), 1, 1)
    end_dt = datetime.date(int(end_year), 12, 31)
    delta_days = (end_dt - start_dt).days
    random_days = random.randint(0, max(0, delta_days))
    d = start_dt + datetime.timedelta(days=random_days)
    return d.strftime(date_format)


def _generate_address(city_filter=None):
    city = city_filter if city_filter in KOREA_DISTRICTS else random.choice(KOREA_CITIES)
    districts = KOREA_DISTRICTS.get(city, ["중앙로 123"])
    dist = random.choice(districts)
    bldg_num = random.randint(1, 250)
    room_num = f"{random.randint(1, 20)}층 {random.randint(101, 2004)}호"
    return f"{city} {dist} {bldg_num} ({room_num})"


def _parse_kv_mapping(mapping_str):
    """'키:값, 키2:값2' 문자열을 딕셔너리로 파싱"""
    mapping = {}
    if not mapping_str:
        return mapping
    pairs = [p.strip() for p in mapping_str.replace('\n', ',').split(',') if p.strip()]
    for p in pairs:
        if ':' in p:
            k, v = p.split(':', 1)
            mapping[k.strip()] = v.strip()
        elif '=' in p:
            k, v = p.split('=', 1)
            mapping[k.strip()] = v.strip()
        elif '->' in p:
            k, v = p.split('->', 1)
            mapping[k.strip()] = v.strip()
    return mapping


def _lookup_kv(mapping, source_val, fallback=""):
    """키-값 매핑에서 정방향(Key->Val) 및 역방향(Val->Key) 검색"""
    if not mapping:
        return fallback
    if not source_val:
        return random.choice(list(mapping.values())) if mapping else fallback
    s_val = str(source_val).strip()
    # 1. 정방향 일치
    if s_val in mapping:
        return mapping[s_val]
    # 2. 역방향 일치
    for k, v in mapping.items():
        if v == s_val:
            return k
    # 3. 대소문자 무시 일치
    for k, v in mapping.items():
        if k.lower() == s_val.lower():
            return v
    return fallback or random.choice(list(mapping.values())) if mapping else s_val


def generate_mock_rows(schema, count=50):
    """
    사용자가 정의한 스키마에 따라 모의 데이터 행(Row) 생성
    - 부서:부서코드, 직급:직급코드 등 연계 키-값(Key-Value) 매핑 지원
    """
    count = max(1, min(int(count), 50000))
    rows = []

    for row_idx in range(count):
        row = {}
        cached_name = None
        deferred_kv_cols = []

        # 1차 패스: 기본 독립 컬럼 생성
        for col in schema:
            col_id = col.get("id", f"col_{random.randint(100, 999)}")
            col_name = col.get("name", "컬럼")
            col_type = col.get("type", "text")
            opts = col.get("options", {})

            if col_type in ("key_value", "linked_map"):
                deferred_kv_cols.append(col)
                continue

            val = ""
            if col_type == "name":
                name, _, _ = _generate_korean_name(opts.get("gender", "any"))
                val = name
                cached_name = name
            elif col_type == "email":
                name = cached_name or _generate_korean_name()[0]
                val = _generate_email(name, opts.get("domains", ""))
            elif col_type == "choices":
                raw_choices = opts.get("choices", "")
                choice_list = [c.strip() for c in raw_choices.split(",") if c.strip()]
                if not choice_list:
                    choice_list = ["기본값 A", "기본값 B", "기본값 C"]
                val = random.choice(choice_list)
            elif col_type == "phone":
                val = _generate_phone(opts.get("format", "010-XXXX-XXXX"))
            elif col_type == "date":
                s_yr = opts.get("start_year", 2020)
                e_yr = opts.get("end_year", 2026)
                fmt = opts.get("format", "%Y-%m-%d")
                val = _generate_date(s_yr, e_yr, fmt)
            elif col_type == "number":
                min_v = int(opts.get("min", 1000))
                max_v = int(opts.get("max", 100000))
                step = int(opts.get("step", 1))
                val = random.randrange(min_v, max(min_v + 1, max_v + 1), max(1, step))
            elif col_type == "sequence":
                prefix = opts.get("prefix", "")
                start_num = int(opts.get("start_num", 1))
                pad = int(opts.get("padding", 4))
                seq = start_num + row_idx
                val = f"{prefix}{str(seq).zfill(pad)}"
            elif col_type == "biz_no":
                val = _generate_biz_no()
            elif col_type == "address":
                val = _generate_address(opts.get("city", ""))
            elif col_type == "company":
                val = random.choice(COMPANY_PRESETS)
            elif col_type == "product":
                prod, price = random.choice(PRODUCTS)
                val = prod
            elif col_type == "uuid":
                import uuid
                val = str(uuid.uuid4())
            else:
                val = f"데이터_{row_idx + 1}"

            row[col_name] = val

        # 2차 패스: 참조 대상 컬럼의 값을 기반으로 연계 키-값(Key-Value) 매핑 생성
        for col in deferred_kv_cols:
            col_name = col.get("name", "연계컬럼")
            opts = col.get("options", {})
            mapping = _parse_kv_mapping(opts.get("mapping", ""))
            target_col = opts.get("target_column", "")
            output_part = opts.get("output_part", "value")  # "value" or "key"

            source_val = row.get(target_col, "") if target_col else ""
            if source_val and mapping:
                val = _lookup_kv(mapping, source_val)
            elif mapping:
                # 참조 컬럼이 비어있거나 매핑이 없으면 랜덤 키/값 선택
                pair = random.choice(list(mapping.items()))
                val = pair[0] if output_part == "key" else pair[1]
                # 참조 컬럼이 정의되어 있고 아직 행에 없다면 동기화
                if target_col and target_col not in row:
                    row[target_col] = pair[1] if output_part == "key" else pair[0]
            else:
                val = f"KV_{row_idx + 1}"

            row[col_name] = val

        rows.append(row)

    return rows


def build_excel_workbook(schema, rows, sheet_name="MockData"):
    """openpyxl을 사용하여 프로페셔널한 스타일 서식이 적용된 Excel 워크북 객체 생성"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = [col.get("name", "컬럼") for col in schema]
    ws.append(headers)

    # 1. 헤더 스타일 (세련된 네이비/인디고 배경 + 흰색 굵은 폰트)
    header_fill = PatternFill(start_color="3730A3", end_color="3730A3", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.row_dimensions[1].height = 28

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # 2. 데이터 행 추가 및 서식
    body_font = Font(name="맑은 고딕", size=10)
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for r_idx, row_dict in enumerate(rows, start=2):
        row_values = [row_dict.get(h, "") for h in headers]
        ws.append(row_values)
        ws.row_dimensions[r_idx].height = 22

        is_even = (r_idx % 2 == 0)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col_num)
            cell.font = body_font
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill

            # 정렬 및 숫자 포맷
            val = cell.value
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            elif isinstance(val, str) and (val.startswith("010-") or val.startswith("EMP-") or len(val) <= 10):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 3. 열 너비 자동 맞춤
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            cell_len = sum(2 if ord(char) > 128 else 1 for char in val_str)
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 4, 50))

    return wb


# ==========================================
# Eel Exposed API 함수들
# ==========================================

@eel.expose
def preview_mock_data(schema, count=10):
    """실시간 화면 미리보기용 모의 데이터 행 생성 (JSON 반환)"""
    try:
        rows = generate_mock_rows(schema, count=min(int(count), 20))
        headers = [col.get("name", "컬럼") for col in schema]
        return {
            "status": "success",
            "headers": headers,
            "rows": rows,
            "total_count": len(rows)
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


@eel.expose
def save_mock_data_excel_dialog(schema, count=100, default_filename="mock_data.xlsx"):
    """Tkinter 파일 저장 대화상자를 열어 지정된 경로에 .xlsx 엑셀 파일 생성 및 저장"""
    try:
        count = int(count)
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)

        now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        if not default_filename.endswith(".xlsx"):
            default_filename = f"{default_filename}_{now_str}.xlsx"

        file_path = filedialog.asksaveasfilename(
            parent=root,
            title="모의 데이터 엑셀(.xlsx) 파일 저장",
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel 통합 문서 (*.xlsx)", "*.xlsx"), ("모든 파일 (*.*)", "*.*")]
        )
        root.destroy()

        if not file_path:
            return {"status": "cancelled"}

        rows = generate_mock_rows(schema, count=count)
        wb = build_excel_workbook(schema, rows, sheet_name="생성데이터")
        wb.save(file_path)

        return {
            "status": "success",
            "file_path": file_path,
            "row_count": len(rows),
            "message": f"총 {len(rows)}건의 데이터가 성공적으로 엑셀 파일로 저장되었습니다."
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


@eel.expose
def save_mock_data_csv_dialog(schema, count=100, default_filename="mock_data.csv"):
    """Tkinter 파일 저장 대화상자를 열어 CSV (UTF-8 with BOM) 저장"""
    try:
        import csv
        count = int(count)
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)

        file_path = filedialog.asksaveasfilename(
            parent=root,
            title="모의 데이터 CSV 파일 저장",
            initialfile=default_filename,
            defaultextension=".csv",
            filetypes=[("CSV (쉼표로 구분) (*.csv)", "*.csv"), ("모든 파일 (*.*)", "*.*")]
        )
        root.destroy()

        if not file_path:
            return {"status": "cancelled"}

        rows = generate_mock_rows(schema, count=count)
        headers = [col.get("name", "컬럼") for col in schema]

        with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for r in rows:
                writer.writerow(r)

        return {
            "status": "success",
            "file_path": file_path,
            "row_count": len(rows),
            "message": f"총 {len(rows)}건의 데이터가 CSV 파일로 저장되었습니다."
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}
